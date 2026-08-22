#!/usr/bin/env python3
"""Migrate class timetable sheets from the Excel workbook into CSV source-of-truth files."""

from __future__ import annotations

import csv
import re
from collections import OrderedDict, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "archive" / "TimeTable - MCE.xlsx"
DATA = ROOT / "data"

PERIOD_COLS = {1: 9, 2: 17, 3: 25, 4: 41, 5: 49, 6: 57}
DAY_ROWS = {5: "MON", 7: "TUE", 9: "WED", 11: "THU", 13: "FRI", 15: "SAT"}

# Excel sheet name → section metadata (college-wide friendly ids)
SHEETS = {
    "CSE 3rd": {
        "id": "CSE-3",
        "program": "CSE",
        "program_name": "Computer Science and Engineering",
        "semester": 3,
        "department": "CSE",
    },
    "CSE AI 3rd": {
        "id": "CSE-AI-3",
        "program": "CSE-AI",
        "program_name": "CSE (Artificial Intelligence)",
        "semester": 3,
        "department": "CSE",
    },
    "CSE 4th": {
        "id": "CSE-4",
        "program": "CSE",
        "program_name": "Computer Science and Engineering",
        "semester": 4,
        "department": "CSE",
    },
    "CSE AI 4th": {
        "id": "CSE-AI-4",
        "program": "CSE-AI",
        "program_name": "CSE (Artificial Intelligence)",
        "semester": 4,
        "department": "CSE",
    },
    "CSE 6th": {
        "id": "CSE-6",
        "program": "CSE",
        "program_name": "Computer Science and Engineering",
        "semester": 6,
        "department": "CSE",
    },
    "CSE AI 6th": {
        "id": "CSE-AI-6",
        "program": "CSE-AI",
        "program_name": "CSE (Artificial Intelligence)",
        "semester": 6,
        "department": "CSE",
    },
}

# Canonical faculty names → stable ids (merge Excel spelling variants)
FACULTY_ALIASES = {
    "mr. chandra shekhar singh chandal": ("chandra-shekhar-singh-chandal", "Mr. Chandra Shekhar Singh Chandal", "ECE"),
    "mr. harshit kumar": ("harshit-kumar", "Mr. Harshit Kumar", "CSE"),
    "mr. md sharique eliyas": ("md-sharique-eliyas", "Mr. Md Sharique Eliyas", "CSE"),
    "mr. ashish kumar": ("ashish-kumar", "Mr. Ashish Kumar", "ASH"),
    "ms. nisha kumari": ("nisha-kumari", "Ms. Nisha Kumari", "CSE"),
    "ms nisha kumari": ("nisha-kumari", "Ms. Nisha Kumari", "CSE"),
    "ms. priyanshu": ("priyanshu-jha", "Ms. Priyanshu Jha", "CSE"),
    "ms priyanshu": ("priyanshu-jha", "Ms. Priyanshu Jha", "CSE"),
    "ms. priyanshu jha": ("priyanshu-jha", "Ms. Priyanshu Jha", "CSE"),
    "mr. ravi kumar": ("ravi-kumar", "Mr. Ravi Kumar", "CSE"),
    "mr. mohammad aknan": ("mohammad-aknan", "Mr. Mohammad Aknan", "CSE"),
    "ms. rashmi priya": ("rashmi-priya", "Ms. Rashmi Priya", "ECE"),
    "mr. hariom kumar": ("hariom-kumar", "Mr. Hariom Kumar", "CSE"),
    "dr. aditya kumar singh": ("aditya-kumar-singh", "Dr. Aditya Kumar Singh", "ASH"),
    "mr. md asif jamal": ("asif-jamal", "Mr. Md Asif Jamal", "CSE"),
    "mr. asif jamal": ("asif-jamal", "Mr. Md Asif Jamal", "CSE"),
    "mr. mukul anand": ("mukul-anand", "Mr. Mukul Anand", "CSE"),
    "mr. shambhu kumar": ("shambhu-kumar", "Mr. Shambhu Kumar", "ASH"),
    "mr. suman patel": ("suman-patel", "Mr. Suman Patel", "CSE"),
    "dr. kahkashan kouser": ("kahkashan-kouser", "Dr. Kahkashan Kouser", "CSE"),
    "ms. juhi kumari": ("juhi-kumari", "Ms. Juhi Kumari", "CSE"),
    "ms juhi kumari": ("juhi-kumari", "Ms. Juhi Kumari", "CSE"),
    "mr. arvind kumar": ("arvind-kumar", "Mr. Arvind Kumar", "CSE"),
}

LAB_ROOM_HINTS = {
    121: "lab",
    122: "lab",
    127: "lab",
    131: "lab",
    134: "lab",
    204: "lab",
}


def norm_space(value: str) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def faculty_from_name(raw: str) -> tuple[str, str, str]:
    key = norm_space(raw).lower().rstrip(".")
    if key in FACULTY_ALIASES:
        return FACULTY_ALIASES[key]
    # fallback slug
    slug = re.sub(r"[^a-z0-9]+", "-", key)
    slug = re.sub(r"^(mr|ms|dr|mrs)-", "", slug).strip("-")
    return slug, norm_space(raw), "CSE"


def clean_subject_code(raw) -> str:
    if raw is None:
        return ""
    if isinstance(raw, float):
        return str(int(raw)) if raw == int(raw) else str(raw)
    text = str(raw).strip()
    text = re.sub(r"\(\+P\)", "", text, flags=re.I)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.strip()


def parse_ltp(raw) -> tuple[str, str, str]:
    text = norm_space(raw) if raw is not None else ""
    if not text or text == "-":
        return "", "", ""
    m = re.match(r"^(\d+)\s*-\s*(\d+)\s*-\s*(\d+)$", text)
    if not m:
        return "", "", ""
    return m.group(1), m.group(2), m.group(3)


def parse_room_token(token: str) -> str:
    token = norm_space(token)
    if not token or set(token) <= {"-", ".", " "}:
        return ""
    m = re.search(r"(\d{2,4})", token)
    return m.group(1) if m else ""


def split_rooms(room_cell) -> list[str]:
    if room_cell is None:
        return []
    text = str(room_cell)
    # "Room No. - 134 / 204" or "Room No - 203"
    text = re.sub(r"(?i)room\s*no\.?\s*-?", "", text)
    parts = [parse_room_token(p) for p in text.split("/")]
    return parts


def parse_slot_fragment(fragment: str) -> dict | None:
    """Parse pieces like 'OOP', 'OOP LAB G1', 'ML(T)', 'DE LAB G2', 'NPTEL G1'."""
    text = norm_space(fragment)
    if not text:
        return None

    group = ""
    gm = re.search(r"\bG([12])\b", text, flags=re.I)
    if gm:
        group = f"G{gm.group(1)}"
        text = re.sub(r"\bG[12]\b", "", text, flags=re.I)

    is_lab = bool(re.search(r"\bLAB\b", text, flags=re.I))
    is_tut = bool(re.search(r"\(T\)", text, flags=re.I))
    text = re.sub(r"\bLAB\b", "", text, flags=re.I)
    text = re.sub(r"\(T\)", "", text, flags=re.I)
    short = norm_space(text).upper()
    short = short.replace(" ", "")
    # normalize I&I etc already fine
    if short in {"I&I", "IANDI"}:
        short = "I&I"

    if is_lab:
        slot_type = "P"
    elif is_tut:
        slot_type = "T"
    else:
        slot_type = "L"

    # NPTEL with group is treated as practical/batch work
    if short == "NPTEL" and group:
        slot_type = "P"

    return {"short": short, "type": slot_type, "group": group}


def split_cell_subjects(subj_cell) -> list[dict]:
    if subj_cell is None:
        return []
    text = norm_space(subj_cell)
    if not text:
        return []
    parts = [p.strip() for p in text.split("/") if p.strip()]
    out = []
    for part in parts:
        parsed = parse_slot_fragment(part)
        if parsed:
            out.append(parsed)
    return out


def extract_meta(ws) -> dict:
    meta = {}
    for c in range(1, 83):
        v = ws.cell(2, c).value
        if not v:
            continue
        text = str(v)
        if "A.Y" in text:
            meta["academic_year"] = text.split(":-")[-1].strip()
        elif "Branch" in text:
            meta["branch_label"] = text.split(":-")[-1].strip()
        elif "Semester" in text:
            sem = text.split(":-")[-1].strip()
            meta["semester_roman"] = sem
        elif "Session" in text:
            meta["batch_session"] = text.split(":-")[-1].strip()
        elif "w.e.f" in text.lower():
            meta["wef"] = text.split(":-")[-1].strip()
    return meta


def extract_subjects(ws) -> list[dict]:
    rows = []
    for r in range(19, 30):
        code_raw = ws.cell(r, 11).value
        if code_raw is None or str(code_raw).strip() == "":
            # sometimes name alone without code
            continue
        code = clean_subject_code(code_raw)
        if not code:
            continue
        name = norm_space(ws.cell(r, 21).value or "")
        short = norm_space(ws.cell(r, 39).value or "").upper()
        l, t, p = parse_ltp(ws.cell(r, 17).value)
        faculty_raw = ws.cell(r, 43).value
        faculty_id = ""
        faculty_name = ""
        if faculty_raw:
            faculty_id, faculty_name, _ = faculty_from_name(str(faculty_raw))
        rows.append(
            {
                "code": code,
                "name": name,
                "short": short,
                "L": l,
                "T": t,
                "P": p,
                "faculty_id": faculty_id,
                "faculty_name": faculty_name,
            }
        )
    return rows


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_meta_yaml(path: Path) -> None:
    content = """# College-wide timetable metadata
college: Motihari College of Engineering, Motihari
college_short: MCE Motihari
timezone: Asia/Kolkata
days:
  - MON
  - TUE
  - WED
  - THU
  - FRI
  - SAT
periods:
  - id: 1
    label: I
    start: "10:00"
    end: "11:00"
  - id: 2
    label: II
    start: "11:00"
    end: "12:00"
  - id: 3
    label: III
    start: "12:00"
    end: "13:00"
  - id: 4
    label: IV
    start: "14:00"
    end: "15:00"
  - id: 5
    label: V
    start: "15:00"
    end: "16:00"
  - id: 6
    label: VI
    start: "16:00"
    end: "17:00"
breaks:
  - after_period: 3
    label: Lunch
    start: "13:00"
    end: "14:00"
notes: |
  Source of truth is data/sections.csv plus data/schedules/<section-id>/.
  Class / faculty / room views, conflicts, LTP and load are derived from current slots of active sections.
  To add another branch or semester: add a row in sections.csv, then schedules/<id>/slots.csv and offerings.csv.
"""
    path.write_text(content, encoding="utf-8")


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    faculties: OrderedDict[str, dict] = OrderedDict()
    rooms: OrderedDict[str, dict] = OrderedDict()
    subjects: OrderedDict[str, dict] = OrderedDict()
    sections: list[dict] = []
    offerings: list[dict] = []
    slots: list[dict] = []
    slot_n = 0

    def ensure_room(room_id: str) -> None:
        if not room_id or room_id in rooms:
            return
        rtype = LAB_ROOM_HINTS.get(int(room_id) if room_id.isdigit() else -1, "classroom")
        rooms[room_id] = {
            "id": room_id,
            "name": f"Room {room_id}",
            "type": rtype,
            "department": "",
            "capacity": "",
        }

    def ensure_faculty(fid: str, name: str, dept: str) -> None:
        if not fid:
            return
        if fid not in faculties:
            faculties[fid] = {
                "id": fid,
                "name": name,
                "department": dept,
                "email": "",
            }

    for sheet_name, section_base in SHEETS.items():
        ws = wb[sheet_name]
        meta = extract_meta(ws)
        section_id = section_base["id"]
        sections.append(
            {
                "id": section_id,
                "program": section_base["program"],
                "program_name": section_base["program_name"],
                "semester": section_base["semester"],
                "department": section_base["department"],
                "batch_session": meta.get("batch_session", ""),
                "academic_year": meta.get("academic_year", ""),
                "wef": meta.get("wef", ""),
                "status": "active",
                "label": f"{section_base['program']} Sem {section_base['semester']}",
            }
        )

        subj_rows = extract_subjects(ws)
        short_to_code: dict[str, str] = {}
        for s in subj_rows:
            short_to_code[s["short"]] = s["code"]
            if s["code"] not in subjects:
                subjects[s["code"]] = {
                    "code": s["code"],
                    "name": s["name"],
                    "short": s["short"],
                    "L": s["L"],
                    "T": s["T"],
                    "P": s["P"],
                    "department": section_base["department"],
                }
            if s["faculty_id"]:
                _, _, dept = faculty_from_name(s["faculty_name"])
                # use alias dept if available
                for alias in FACULTY_ALIASES.values():
                    if alias[0] == s["faculty_id"]:
                        dept = alias[2]
                        break
                ensure_faculty(s["faculty_id"], s["faculty_name"], dept)
                offerings.append(
                    {
                        "section_id": section_id,
                        "subject_code": s["code"],
                        "faculty_id": s["faculty_id"],
                    }
                )

        for day_row, day in DAY_ROWS.items():
            room_row = day_row + 1
            for period, col in PERIOD_COLS.items():
                subj_cell = ws.cell(day_row, col).value
                room_cell = ws.cell(room_row, col).value
                fragments = split_cell_subjects(subj_cell)
                if not fragments:
                    continue
                room_parts = split_rooms(room_cell)
                # pad rooms to match fragments
                while len(room_parts) < len(fragments):
                    room_parts.append("")

                for i, frag in enumerate(fragments):
                    short = frag["short"]
                    code = short_to_code.get(short, "")
                    if not code:
                        # try without odd chars
                        code = short_to_code.get(short.replace("&", ""), "")
                    faculty_id = ""
                    for off in offerings:
                        if off["section_id"] == section_id and off["subject_code"] == code:
                            faculty_id = off["faculty_id"]
                            break
                    room_id = room_parts[i] if i < len(room_parts) else ""
                    ensure_room(room_id)
                    slot_n += 1
                    notes = ""
                    if not code:
                        notes = f"unresolved short={short}"
                    slots.append(
                        {
                            "id": f"S{slot_n:04d}",
                            "section_id": section_id,
                            "day": day,
                            "period": period,
                            "hours": 1,
                            "subject_code": code,
                            "subject_short": short,
                            "type": frag["type"],
                            "group": frag["group"],
                            "room_id": room_id,
                            "faculty_id": faculty_id,
                            "notes": notes,
                        }
                    )

    # Seed known rooms even if unused in slots
    for rid in ["121", "122", "123", "125", "127", "131", "134", "135", "203", "204"]:
        ensure_room(rid)

    write_meta_yaml(DATA / "meta.yaml")
    write_csv(
        DATA / "sections.csv",
        [
            "id",
            "program",
            "program_name",
            "semester",
            "department",
            "batch_session",
            "academic_year",
            "wef",
            "status",
            "label",
        ],
        sections,
    )
    write_csv(
        DATA / "faculties.csv",
        ["id", "name", "department", "email"],
        list(faculties.values()),
    )
    write_csv(
        DATA / "rooms.csv",
        ["id", "name", "type", "department", "capacity"],
        list(rooms.values()),
    )
    write_csv(
        DATA / "subjects.csv",
        ["code", "name", "short", "L", "T", "P", "department"],
        list(subjects.values()),
    )
    slot_fields = [
        "id",
        "section_id",
        "day",
        "period",
        "hours",
        "subject_code",
        "subject_short",
        "type",
        "group",
        "room_id",
        "faculty_id",
        "notes",
    ]
    slots_by: dict[str, list] = defaultdict(list)
    offerings_by: dict[str, list] = defaultdict(list)
    for row in slots:
        slots_by[row["section_id"]].append(row)
    for row in offerings:
        offerings_by[row["section_id"]].append(row)
    for section in sections:
        folder = DATA / "schedules" / section["id"]
        folder.mkdir(parents=True, exist_ok=True)
        write_csv(folder / "slots.csv", slot_fields, slots_by[section["id"]])
        write_csv(
            folder / "offerings.csv",
            ["section_id", "subject_code", "faculty_id"],
            offerings_by[section["id"]],
        )

    unresolved = [s for s in slots if s["notes"]]
    print(f"Wrote {len(sections)} sections, {len(subjects)} subjects, {len(faculties)} faculties,")
    print(f"      {len(rooms)} rooms, {len(offerings)} offerings, {len(slots)} slots")
    print(f"Unresolved short→code: {len(unresolved)}")
    for u in unresolved[:20]:
        print(" ", u)


if __name__ == "__main__":
    main()
